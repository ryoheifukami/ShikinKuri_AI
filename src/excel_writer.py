"""資金繰り表をExcelで出力する。

- シート1: 資金繰り表(実績+予測を縦に並べる) — 単位: 千円
- シート2: 経営アドバイス
- シート3: 前提・注意事項
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .advisor import AdviceItem
from .financial_model import ROW_LABELS, CashFlowResult, PeriodRow


UNIT_DIVISOR = 1000  # 円 → 千円
UNIT_LABEL = "千円"

HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
ACTUAL_FILL = PatternFill("solid", fgColor="FFFFFF")
FORECAST_FILL = PatternFill("solid", fgColor="FCE4D6")  # 予測=薄オレンジ
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _scaled(v: float | None) -> float | None:
    if v is None:
        return None
    return round(v / UNIT_DIVISOR, 0)


def _value_for(row: PeriodRow, key: str) -> float | None:
    if key.startswith("__section"):
        return None
    return _scaled(getattr(row, key, None))


def _write_table(
    ws,
    title: str,
    periods: list[PeriodRow],
    *,
    start_row: int,
) -> int:
    """表を書いて、次の開始行を返す。"""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    r = start_row + 1

    # ヘッダー行
    ws.cell(row=r, column=1, value="項目").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    for i, p in enumerate(periods, start=2):
        header_text = p.period
        if p.label:
            header_text = f"{p.label}\n{p.period}"
        c = ws.cell(row=r, column=i, value=header_text)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    for key, label in ROW_LABELS:
        cell_label = ws.cell(row=r, column=1, value=label)
        cell_label.border = THIN_BORDER

        if key.startswith("__section"):
            cell_label.fill = SECTION_FILL
            cell_label.font = Font(bold=True)
            for i in range(2, 2 + len(periods)):
                c = ws.cell(row=r, column=i)
                c.fill = SECTION_FILL
                c.border = THIN_BORDER
            r += 1
            continue

        bold = key in ("opening_cash", "ending_cash", "net_change", "operating_income", "gross_profit")
        if bold:
            cell_label.font = Font(bold=True)

        for i, p in enumerate(periods, start=2):
            val = _value_for(p, key)
            c = ws.cell(row=r, column=i, value=val)
            c.number_format = "#,##0;[Red]-#,##0"
            c.border = THIN_BORDER
            if p.is_forecast:
                c.fill = FORECAST_FILL
            if bold:
                c.font = Font(bold=True)
        r += 1

    ws.column_dimensions["A"].width = 22
    for i in range(2, 2 + len(periods)):
        ws.column_dimensions[get_column_letter(i)].width = 18

    return r + 1


def _write_advice_sheet(ws, advice: list[AdviceItem]) -> None:
    ws.cell(row=1, column=1, value="経営アドバイス").font = Font(bold=True, size=14)
    headers = ["観点", "優先度", "見出し", "詳細", "推奨アクション"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

    priority_label = {"high": "高", "medium": "中", "low": "低"}
    priority_color = {"high": "F8CBAD", "medium": "FFF2CC", "low": "E2EFDA"}
    category_color = {
        "資金繰り": "D9E1F2",
        "安全性": "FCE4D6",
        "収益性": "E2EFDA",
        "成長性": "FFF2CC",
        "経営課題": "EDEDED",
    }
    for i, a in enumerate(advice, start=4):
        c_cat = ws.cell(row=i, column=1, value=a.category)
        c_cat.fill = PatternFill("solid", fgColor=category_color.get(a.category, "FFFFFF"))
        c_cat.font = Font(bold=True)
        c_cat.alignment = Alignment(horizontal="center", vertical="center")
        c_cat.border = THIN_BORDER

        c_p = ws.cell(row=i, column=2, value=priority_label.get(a.priority, a.priority))
        c_p.fill = PatternFill("solid", fgColor=priority_color.get(a.priority, "FFFFFF"))
        c_p.alignment = Alignment(horizontal="center", vertical="center")
        c_p.border = THIN_BORDER

        c_h = ws.cell(row=i, column=3, value=a.headline)
        c_h.font = Font(bold=True)
        c_h.alignment = Alignment(wrap_text=True, vertical="top")
        c_h.border = THIN_BORDER

        c_d = ws.cell(row=i, column=4, value=a.detail)
        c_d.alignment = Alignment(wrap_text=True, vertical="top")
        c_d.border = THIN_BORDER

        c_a = ws.cell(row=i, column=5, value=a.action)
        c_a.alignment = Alignment(wrap_text=True, vertical="top")
        c_a.font = Font(color="305496")
        c_a.border = THIN_BORDER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 50


def _write_notes_sheet(ws, actual: CashFlowResult, user_notes: str | None = None) -> None:
    ws.cell(row=1, column=1, value="前提・注意事項").font = Font(bold=True, size=14)

    r = 3
    ws.cell(row=r, column=1, value=f"モード: {'月次' if actual.mode == 'monthly' else '年次'}")
    r += 1
    if actual.company_name:
        ws.cell(row=r, column=1, value=f"会社名: {actual.company_name}")
        r += 1
    if actual.fiscal_year_end:
        ws.cell(row=r, column=1, value=f"決算日: {actual.fiscal_year_end}")
        r += 1
    r += 2

    # ユーザー追記情報
    if user_notes and user_notes.strip():
        ws.cell(row=r, column=1, value="追記情報(担当者記入)").font = Font(bold=True, size=12, color="305496")
        r += 1
        for line in user_notes.strip().splitlines():
            c = ws.cell(row=r, column=1, value=line)
            c.alignment = Alignment(wrap_text=True)
            r += 1
        r += 1

    if actual.notes:
        ws.cell(row=r, column=1, value="データ注記").font = Font(bold=True)
        r += 1
        for n in actual.notes:
            c = ws.cell(row=r, column=1, value=f"• {n}")
            c.alignment = Alignment(wrap_text=True)
            r += 1

    if actual.warnings:
        r += 1
        ws.cell(row=r, column=1, value="警告").font = Font(bold=True, color="C00000")
        r += 1
        for w in actual.warnings:
            c = ws.cell(row=r, column=1, value=f"• {w}")
            c.alignment = Alignment(wrap_text=True)
            r += 1

    ws.column_dimensions["A"].width = 90


def build_workbook(
    actual: CashFlowResult,
    forecasts: list[PeriodRow] | None,
    advice: Iterable[AdviceItem],
    company_name: str | None = None,
    user_notes: str | None = None,
) -> bytes:
    """Excel全体を組み立てて、xlsxのバイト列を返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "資金繰り表"

    title = "資金繰り表"
    if company_name:
        title += f"({company_name})"
    title += f"  (単位: {UNIT_LABEL})"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value="オレンジセル: 将来予測 (前期⇄当期の成長率から自動算出)").font = Font(
        italic=True, color="808080"
    )

    # 実績 + 予測を1つの表に並べて出す
    combined = list(actual.periods) + list(forecasts or [])
    if combined:
        section_title = "月次推移" if actual.mode == "monthly" else "年次推移"
        _write_table(ws, section_title, combined, start_row=4)

    ws_advice = wb.create_sheet("経営アドバイス")
    _write_advice_sheet(ws_advice, list(advice))

    ws_notes = wb.create_sheet("前提・注意事項")
    _write_notes_sheet(ws_notes, actual, user_notes=user_notes)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
